__author__ = 'Andrew'
_project_ = 'CIT220a'
# Date: 2016-02-10
# Session 5


import numpy as np

import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference



from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

# Sets up a workbook in excel
wb = Workbook(write_only=True)
ws = wb.create_sheet()

# This creates/names a new sheet, and sets the active sheet
ws = wb.create_sheet(title="Membership Ratio")


# Data - since I lack big data to accurately cover this ratio nicely, just placing in dummy numbers:

rows = [
    ('Month', 'Joined', 'Left'),
    (2, 50, 10),
    (3, 22, 5),
    (4, 21, 7),
    (5, 45, 1),
    (6, 15, 4),
    (7, 12, 3),
]


for row in rows:
    ws.append(row)

# This is supplied sample code to get the export correct for the excel program
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Membership'
chart1.x_axis.title = 'Month'

data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "A10")

from copy import deepcopy

chart2 = deepcopy(chart1)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horizontal Bar Chart"

ws.add_chart(chart2, "G10")


chart3 = deepcopy(chart1)
chart3.type = "col"
chart3.style = 12
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = 'Stacked Chart'

ws.add_chart(chart3, "A27")


chart4 = deepcopy(chart1)
chart4.type = "bar"
chart4.style = 13
chart4.grouping = "percentStacked"
chart4.overlap = 100
chart4.title = 'Percent Stacked Chart'

ws.add_chart(chart4, "G27")

#wb.save("bar.xlsx")

### Pie Chart
import s5_church

from openpyxl import Workbook

from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint

# Numbers - Grab data from the master church program
# Even though this is not "big data" this simulates the ratio of numbers based on current available data
paidStaff = s5_church.paidStaff()
volunteerStaff = s5_church.volunteerStaff()

data = [
    ['Category', 'Number'],
    ['Paid', paidStaff],
    ['Volunteer', volunteerStaff],
]

# We are setting a new tab for this information
ws = wb.create_sheet(title="Ministry Paid")


for row in data:
    ws.append(row)

pie = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=5)
data = Reference(ws, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Volunteers vs. Paid"

# Cut the first slice out of the pie
slice = DataPoint(idx=0, explosion=20)
pie.series[0].data_points = [slice]

ws.add_chart(pie, "D1")

### Line Chart
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis


ws = wb.create_sheet(title="Attendance")
# Attendance Numbers - no big data equivalent in my sample, but something that is ideal for a church to track
# Example of how this should be used.

rows1 = [
    ['Date', 'Total Count', 'Saturday', 'Sunday'],
    [date(2015,9, 1), 550, 100, 450],
    [date(2015,9, 8), 400, 75, 325],
    [date(2015,9, 15), 550, 125, 325],
    [date(2015,9, 22), 700, 150, 550],
    [date(2015,9, 29), 600, 100, 500],
    ]

for row1 in rows1:
    ws.append(row1)

c1 = LineChart()
c1.title = "Line Chart"
c1.style = 13
c1.y_axis.title = 'Number of People Counted'
c1.x_axis.title = 'Service week'

data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
c1.add_data(data, titles_from_data=True)

# Style the lines
s1 = c1.series[0]
s1.marker.symbol = "triangle"
s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

s1.graphicalProperties.line.noFill = True

s2 = c1.series[1]
s2.graphicalProperties.line.solidFill = "00AAAA"
s2.graphicalProperties.line.dashStyle = "sysDot"
s2.graphicalProperties.line.width = 100050 # width in EMUs

s2 = c1.series[2]
s2.smooth = True # Make the line smooth

ws.add_chart(c1, "A10")

from copy import deepcopy
stacked = deepcopy(c1)
stacked.grouping = "stacked"
stacked.title = "Stacked Line Chart"
ws.add_chart(stacked, "A27")

percent_stacked = deepcopy(c1)
percent_stacked.grouping = "percentStacked"
percent_stacked.title = "Percent Stacked Line Chart"
ws.add_chart(percent_stacked, "A44")

# Chart with date axis
c2 = LineChart()
c2.title = "Date Axis"
c2.style = 12
c2.y_axis.title = "Size"
c2.y_axis.crossAx = 500
c2.x_axis = DateAxis(crossAx=100)
c2.x_axis.number_format = 'd-mmm'
c2.x_axis.majorTimeUnit = "days"
c2.x_axis.title = "Date"

c2.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=7)
c2.set_categories(dates)

ws.add_chart(c2, "A61")


wb.save("Church_Report.xlsx")












